from io import BytesIO

from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

KARAMS_ORANGE = 'FFFF9220'
HEADER_FONT = Font(bold=True, color='FFFFFFFF')


def _header_row(ws, row, headers):
    fill = PatternFill(start_color=KARAMS_ORANGE, end_color=KARAMS_ORANGE, fill_type='solid')
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')


def gerar_xlsx_produtividade(context):
    wb = Workbook()
    ws = wb.active
    ws.title = 'KPIs'

    filtros = context['filtros']
    ws.cell(row=1, column=1, value='Produtividade Comercial — CRM Karams')
    ws.cell(row=2, column=1, value=f"Período: {filtros['de']} a {filtros['ate']}")

    _header_row(ws, 4, ['Indicador', 'Valor'])
    realizado = context['realizado']
    conversao = context['conversao']
    conversao_orc = context['conversao_orcamentos']
    rows = [
        ('Total de Contatos', realizado['contatos']),
        ('Clientes Novos', realizado['clientes_novos']),
        ('Total Interações', realizado['total_interacoes']),
        ('Propostas / Orçamentos', realizado['propostas']),
        ('Conversão Prospecção (%)', conversao['taxa_pct']),
        ('Conversão Orçamentos (%)', conversao_orc['taxa_pct']),
        ('Vendas (R$)', float(realizado['vendas_valor'])),
    ]
    for i, (label, val) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    ws2 = wb.create_sheet('Sem contato 30d')
    _header_row(ws2, 1, ['Cliente', 'Último contato', 'Dias sem contato', 'Vendedor'])
    for i, item in enumerate(context['clientes_sem_contato'], start=2):
        ultima = item.get('data_ultimo_contato')
        ws2.cell(row=i, column=1, value=item['cliente'].nome)
        ws2.cell(row=i, column=2, value=ultima.strftime('%d/%m/%Y') if ultima else 'Nunca')
        ws2.cell(row=i, column=3, value=item['dias_sem_contato'])
        v = item['cliente'].vendedor
        ws2.cell(row=i, column=4, value=v.get_full_name() if v else '—')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_pdf_produtividade(context):
    from xhtml2pdf import pisa

    html = render_to_string('relatorios/produtividade_pdf.html', context)
    buffer = BytesIO()
    pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    buffer.seek(0)
    return buffer
